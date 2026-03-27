#!/usr/bin/env python3
"""
Download O*NET database and extract required files.

O*NET releases ~4x/year (Feb, May, Aug, Nov).
URL pattern: https://www.onetcenter.org/dl_files/database/db_{VV}_{N}_excel.zip
  e.g. db_30_2_excel.zip for O*NET 30.2

Usage:
    python3 scripts/download_onet.py             # download latest known version
    python3 scripts/download_onet.py --version 30.2
    python3 scripts/download_onet.py --check     # check what version is current on site
    python3 scripts/download_onet.py --sync      # update All_Occupations_ONET.csv with new/removed codes

Current version: O*NET 30.2 (February 2026)
"""

import argparse
import csv
import shutil
import sys
import tempfile
import urllib.request
import zipfile
from pathlib import Path

import openpyxl

# ── Config ─────────────────────────────────────────────────────────────────────

ONET_DB_DIR      = Path(__file__).parent.parent / "data" / "input" / "onet_db"
BACKUP_DIR       = Path(__file__).parent.parent / "data" / "input" / "onet_db_backup"
OCCUPATION_CSV   = Path(__file__).parent.parent / "data" / "input" / "All_Occupations_ONET.csv"
SCORES_CSV       = Path(__file__).parent.parent / "data" / "output" / "ai_resilience_scores.csv"

CURRENT_VERSION = "30.2"

# Files we need from the zip
REQUIRED_FILES = [
    "Occupation Data.xlsx",
    "Sample of Reported Titles.xlsx",
    "Education, Training, and Experience.xlsx",
    "Education, Training, and Experience Categories.xlsx",
    "Task Statements.xlsx",
    "Task Ratings.xlsx",
]

# Local names (some O*NET files have slightly different names across versions)
LOCAL_NAMES = {
    "Education, Training, and Experience.xlsx":            "Education Training and Experience.xlsx",
    "Education, Training, and Experience Categories.xlsx": "ETE Categories.xlsx",
}


def version_to_url(version: str) -> str:
    """Convert 'XX.Y' version string to O*NET zip URL."""
    major, minor = version.split(".")
    return f"https://www.onetcenter.org/dl_files/database/db_{major}_{minor}_excel.zip"


def version_to_job_zones_url(version: str) -> str:
    major, minor = version.split(".")
    return f"https://www.onetcenter.org/dl_files/database/db_{major}_{minor}_excel/Job%20Zones.xlsx"


def check_latest_version() -> str | None:
    """Try a few known recent versions to find the latest available."""
    candidates = ["30.3", "30.2", "30.1", "29.3", "29.2"]
    for v in candidates:
        url = version_to_url(v)
        try:
            req = urllib.request.Request(url, method="HEAD")
            with urllib.request.urlopen(req, timeout=10) as resp:
                if resp.status == 200:
                    return v
        except Exception:
            continue
    return None


def backup_existing():
    """Back up current onet_db files to onet_db_backup/."""
    if not ONET_DB_DIR.exists():
        return
    existing = list(ONET_DB_DIR.glob("*.xlsx"))
    if not existing:
        return

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Backing up {len(existing)} files → {BACKUP_DIR}")
    for f in existing:
        dest = BACKUP_DIR / f.name
        shutil.copy2(f, dest)
        print(f"  ✓ {f.name}")


def download_and_extract(version: str, dry_run: bool = False):
    url = version_to_url(version)
    print(f"\nO*NET {version}")
    print(f"URL: {url}")

    if dry_run:
        print("(dry run — not downloading)")
        return

    ONET_DB_DIR.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmpdir:
        zip_path = Path(tmpdir) / "onet.zip"

        print(f"Downloading...", end=" ", flush=True)
        try:
            urllib.request.urlretrieve(url, zip_path)
        except Exception as e:
            print(f"\n✗ Download failed: {e}")
            sys.exit(1)

        size_mb = zip_path.stat().st_size / (1024 * 1024)
        print(f"done ({size_mb:.0f} MB)")

        print("Extracting...")
        with zipfile.ZipFile(zip_path) as zf:
            all_names = zf.namelist()

            for required in REQUIRED_FILES:
                # Find matching entry (may be in a subdirectory)
                match = next((n for n in all_names if n.endswith(required)), None)
                if not match:
                    print(f"  ⚠ Not found in zip: {required}")
                    continue

                local_name = LOCAL_NAMES.get(required, required)
                dest = ONET_DB_DIR / local_name

                with zf.open(match) as src, open(dest, "wb") as dst:
                    shutil.copyfileobj(src, dst)

                size_kb = dest.stat().st_size // 1024
                print(f"  ✓ {local_name} ({size_kb:,} KB)")

    print(f"\nFiles in {ONET_DB_DIR}:")
    for f in sorted(ONET_DB_DIR.glob("*.xlsx")):
        size_kb = f.stat().st_size // 1024
        print(f"  {f.name} ({size_kb:,} KB)")


# ── Sync occupation list ───────────────────────────────────────────────────────

def load_job_zones(version: str) -> dict:
    """Download Job Zones.xlsx and return {code: job_zone_int}."""
    url = version_to_job_zones_url(version)
    print(f"  Downloading Job Zones from {url}...", end=" ", flush=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        urllib.request.urlretrieve(url, tmp_path)
        print("done")
    except Exception as e:
        print(f"✗ {e}")
        return {}

    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: O*NET-SOC Code, Title, Job Zone
        if row[0] and row[2] is not None:
            try:
                result[row[0]] = int(row[2])
            except (TypeError, ValueError):
                pass
    wb.close()
    tmp_path.unlink(missing_ok=True)
    return result


def sync_occupation_list(version: str):
    """
    Diff the current All_Occupations_ONET.csv against the new O*NET Occupation Data.xlsx.
    - New codes: added to the CSV
    - Removed codes: kept as-is (scores preserved; enrich falls back to cache)
    - Reports a summary of changes
    """
    if not OCCUPATION_CSV.exists():
        print(f"✗ {OCCUPATION_CSV} not found")
        sys.exit(1)

    occ_data_path = ONET_DB_DIR / "Occupation Data.xlsx"
    if not occ_data_path.exists():
        print(f"✗ {occ_data_path} not found — run download first")
        sys.exit(1)

    # Load current CSV
    with open(OCCUPATION_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        current_rows = list(reader)
    current_codes = {r["Code"]: r for r in current_rows}

    # Load new O*NET occupation data
    wb = openpyxl.load_workbook(occ_data_path, read_only=True, data_only=True)
    ws = wb.active
    new_onet = {}  # code -> title
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, title = row[0], row[1]
        if code and title:
            new_onet[code] = title
    wb.close()

    added   = sorted(set(new_onet) - set(current_codes))
    removed = sorted(set(current_codes) - set(new_onet))

    # Cross-reference with scored occupations
    scored_codes = set()
    if SCORES_CSV.exists():
        with open(SCORES_CSV, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                scored_codes.add(r["Code"])

    print(f"\nO*NET occupation list diff:")
    print(f"  Current CSV: {len(current_codes)} codes")
    print(f"  New O*NET:   {len(new_onet)} codes")
    print(f"  Added:       {len(added)}")
    print(f"  Removed:     {len(removed)}")

    if removed:
        scored_removed = [c for c in removed if c in scored_codes]
        print(f"\nRemoved codes ({len(removed)}) — kept in CSV, scores preserved:")
        for c in removed:
            flag = " ⚠ has score" if c in scored_codes else ""
            print(f"  {c}  {current_codes[c]['Occupation']}{flag}")
        if scored_removed:
            print(f"\n  ⚠ {len(scored_removed)} removed codes have existing scores.")
            print("  Their enrichment will fall back to the scrape cache.")
            print("  Consider whether to keep or retire them from scoring.")

    if not added:
        print("\nNo new codes to add.")
        return

    # Load Job Zones for new codes
    print(f"\nFetching Job Zones for {len(added)} new codes...")
    job_zones = load_job_zones(version)

    new_rows = []
    missing_jz = []
    for code in added:
        jz = job_zones.get(code)
        if jz is None:
            missing_jz.append(code)
            jz_str = ""
        else:
            jz_str = str(jz)
        new_rows.append({
            "Job Zone":   jz_str,
            "Code":       code,
            "Occupation": new_onet[code],
            "Data-level": "Y",
            "url":        f"https://www.onetonline.org/link/summary/{code}",
        })

    if missing_jz:
        print(f"  ⚠ No Job Zone found for {len(missing_jz)} codes:")
        for c in missing_jz:
            print(f"    {c}  {new_onet[c]}")
        print("  These will have empty Job Zone and be excluded from scoring until fixed.")

    # Append new rows to CSV
    all_rows = current_rows + new_rows
    with open(OCCUPATION_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\n  ✓ Added {len(new_rows)} new codes to {OCCUPATION_CSV}")
    print(f"  Next: python3 scripts/enrich_onet.py  (will scrape new codes)")
    print(f"        python3 scripts/score_occupations.py  (will score new codes)")


# ── CLI ────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Download O*NET database files")
    parser.add_argument("--version",   default=CURRENT_VERSION, help="O*NET version (e.g. 30.2)")
    parser.add_argument("--check",     action="store_true", help="Check latest available version")
    parser.add_argument("--sync",      action="store_true", help="Sync All_Occupations_ONET.csv with new O*NET codes")
    parser.add_argument("--no-backup", action="store_true", help="Skip backup of existing files")
    args = parser.parse_args()

    if args.check:
        print("Checking latest available O*NET version...")
        latest = check_latest_version()
        if latest:
            print(f"✓ Latest available: O*NET {latest}")
            if latest != CURRENT_VERSION:
                print(f"  (current in script: {CURRENT_VERSION} — update CURRENT_VERSION if you want to download)")
        else:
            print("✗ Could not determine latest version")
        return

    if args.sync:
        sync_occupation_list(args.version)
        return

    if not args.no_backup:
        backup_existing()

    download_and_extract(args.version)
    print("\n✓ Done")


if __name__ == "__main__":
    main()
